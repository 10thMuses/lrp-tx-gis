"""
EIA-860 refresh — produces eia860_plants_<date>.csv and eia860_battery_<date>.csv.

Source: https://www.eia.gov/electricity/data/eia860/
Latest published release: 2024-data zip (`eia8602024.zip`), released 2025.
Re-running with `--year 2025` once the 2025-data release lands is a one-arg change.

ARCHITECTURE.md §188 documents that EIA hot-links 503 without a Referer header,
so all fetches send both `User-Agent` and `Referer`.

Per ARCHITECTURE.md §84: capacity / technology / fuel live in the Generator
sheet, NOT the Plant sheet. We aggregate generators by Plant Code with
`Status == 'OP'`, sum Nameplate MW, take mode of Technology + Energy Source 1.
Battery storage lives in its own sheet (`3_3_Energy_Storage_Y<year>.xlsx`).

Schema match: outputs follow `combined_points.csv` 31-column layout. The two
output CSVs are filtered to Texas only (`State == 'TX'`).

Usage:
    python scripts/refresh_eia860.py [--year 2024] [--out outputs/refresh]

Exit codes:
    0  both CSVs written non-empty
    1  fetch failed (logged FETCH_FAILED, no partial output)
    2  unzip / sheet-locate failure
"""
import argparse
import csv
import glob
import io
import os
import re
import sys
import tempfile
import urllib.request
import zipfile
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

import openpyxl

UA = "LRP-TX-GIS/1.0 (refinement-eia860-refresh)"
REFERER = "https://www.eia.gov/electricity/data/eia860/"

# Try canonical path first, then archive/ subpath. EIA shifts files between the
# two without notice when a year transitions from "current" to "archive".
URL_PATTERNS = [
    "https://www.eia.gov/electricity/data/eia860/xls/eia860{year}.zip",
    "https://www.eia.gov/electricity/data/eia860/archive/xls/eia860{year}.zip",
]

# combined_points.csv schema — 31 columns, order is canonical.
SCHEMA = [
    "layer_id", "lat", "lon", "name", "plant_code", "county", "technology",
    "capacity", "sector", "inr", "fuel", "mw", "zone", "poi", "entity",
    "funnel_stage", "group", "under_construction", "commissioned",
    "capacity_mw", "operator", "voltage", "osm_id", "depth_ft", "use",
    "aquifer", "project", "manu", "model", "cap_kw", "year",
]

# Sheet-name patterns inside the EIA-860 zip. Names drift slightly year-to-year
# (`2___Plant_Y2024.xlsx` vs `2___Plant_Y2023.xlsx`); glob handles it.
PLANT_GLOB = "2___Plant_Y*.xlsx"
GEN_GLOB = "3_1_Generator_Y*.xlsx"
BATTERY_GLOBS = [
    "3_3_Energy_Storage_Y*.xlsx",  # 2022+
    "3_3_Battery_Y*.xlsx",         # legacy fallback
]


def fetch_with_retry(url, attempts=5, sleep=10):
    """Standard refresh-cycle fetch helper (OPERATING.md §8). Returns bytes or
    raises last exception."""
    import time
    last = None
    for i in range(attempts):
        try:
            req = urllib.request.Request(
                url, headers={"User-Agent": UA, "Referer": REFERER},
            )
            with urllib.request.urlopen(req, timeout=120) as r:
                return r.read()
        except Exception as e:
            last = e
            sys.stderr.write(f"[fetch-retry {i+1}/{attempts}] {url}: {e}\n")
            time.sleep(sleep)
    raise last


def fetch_eia860_zip(year):
    for pattern in URL_PATTERNS:
        url = pattern.format(year=year)
        try:
            blob = fetch_with_retry(url)
            sys.stderr.write(f"[fetched] {url} ({len(blob)} bytes)\n")
            return blob, url
        except Exception as e:
            sys.stderr.write(f"[skip] {url}: {e}\n")
    raise RuntimeError(f"FETCH_FAILED: no working URL for eia860 year={year}")


def find_sheet(extracted_dir, pattern):
    matches = glob.glob(os.path.join(extracted_dir, pattern))
    if not matches:
        # EIA sometimes nests in a year-named subdir.
        matches = glob.glob(os.path.join(extracted_dir, "**", pattern), recursive=True)
    if not matches:
        return None
    return matches[0]


def detect_header_row(ws, expected_cells):
    """EIA xlsx files put a banner in row 1 and the column header in row 2 or 3.
    Scan first 6 rows; return the 1-indexed row whose values include all of
    `expected_cells` (case-insensitive substring)."""
    expected_lower = [e.lower() for e in expected_cells]
    for r in range(1, 7):
        row_vals = [str(c.value).lower() if c.value is not None else "" for c in ws[r]]
        if all(any(e in v for v in row_vals) for e in expected_lower):
            return r
    return None


def load_xlsx_rows(path, expected_cells):
    """Load an EIA xlsx sheet into list-of-dicts using its detected header row.
    Lazy-loads via openpyxl read_only mode to keep memory bounded. Strings
    only — caller coerces types."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    # read_only mode doesn't allow ws[r] indexing on some versions; convert to list.
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not all_rows:
        return []
    # Locate header row.
    header_idx = None
    expected_lower = [e.lower() for e in expected_cells]
    for r_idx, row in enumerate(all_rows[:6]):
        row_vals = ["" if v is None else str(v).lower() for v in row]
        if all(any(e in v for v in row_vals) for e in expected_lower):
            header_idx = r_idx
            break
    if header_idx is None:
        raise RuntimeError(
            f"could not locate header row in {path} (expected: {expected_cells})"
        )
    header = ["" if v is None else str(v).strip() for v in all_rows[header_idx]]
    out = []
    for row in all_rows[header_idx + 1:]:
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        d = {h: row[i] if i < len(row) else None for i, h in enumerate(header)}
        out.append(d)
    return out


def column_lookup(rows, *candidates):
    """Return the first column name from a row dict that matches any of the
    candidate substrings (case-insensitive). EIA renames columns
    year-over-year (e.g., 'Energy Source 1' vs 'Energy Source Code 1')."""
    if not rows:
        return None
    keys_lower = {k.lower(): k for k in rows[0].keys() if k}
    for cand in candidates:
        cl = cand.lower()
        for kl, k in keys_lower.items():
            if cl in kl:
                return k
    return None


def aggregate_generators(gen_rows):
    """Group generators by Plant Code, filter Status='OP'. Returns dict
    plant_code -> {capacity_mw, technology, fuel, commissioned}."""
    plant_col = column_lookup(gen_rows, "Plant Code")
    status_col = column_lookup(gen_rows, "Status")
    cap_col = column_lookup(gen_rows, "Nameplate Capacity (MW)", "Nameplate Capacity")
    tech_col = column_lookup(gen_rows, "Technology")
    fuel_col = column_lookup(gen_rows, "Energy Source 1", "Energy Source Code 1")
    op_year_col = column_lookup(gen_rows, "Operating Year")

    buckets = defaultdict(lambda: {"caps": [], "techs": [], "fuels": [], "years": []})
    for r in gen_rows:
        status = ("" if r.get(status_col) is None else str(r.get(status_col))).strip().upper()
        if status != "OP":
            continue
        pc = r.get(plant_col)
        if pc is None or str(pc).strip() == "":
            continue
        pc = str(pc).strip()
        try:
            cap = float(r.get(cap_col)) if r.get(cap_col) not in (None, "") else None
        except (TypeError, ValueError):
            cap = None
        if cap is not None:
            buckets[pc]["caps"].append(cap)
        if r.get(tech_col):
            buckets[pc]["techs"].append(str(r.get(tech_col)).strip())
        if r.get(fuel_col):
            buckets[pc]["fuels"].append(str(r.get(fuel_col)).strip())
        if r.get(op_year_col) not in (None, ""):
            try:
                buckets[pc]["years"].append(int(r.get(op_year_col)))
            except (TypeError, ValueError):
                pass

    out = {}
    for pc, agg in buckets.items():
        out[pc] = {
            "capacity_mw": round(sum(agg["caps"]), 1) if agg["caps"] else "",
            "technology": Counter(agg["techs"]).most_common(1)[0][0] if agg["techs"] else "",
            "fuel": Counter(agg["fuels"]).most_common(1)[0][0] if agg["fuels"] else "",
            "commissioned": str(min(agg["years"])) if agg["years"] else "",
        }
    return out


def empty_row():
    return {col: "" for col in SCHEMA}


def write_csv(path, rows):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=SCHEMA)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--year", type=int, default=2024,
                    help="EIA-860 data year (zip is named eia860<year>.zip)")
    ap.add_argument("--out", default="outputs/refresh",
                    help="output directory")
    args = ap.parse_args()

    today = date.today().isoformat()

    try:
        zip_bytes, src_url = fetch_eia860_zip(args.year)
    except Exception as e:
        sys.stderr.write(f"FETCH_FAILED: {e}\n")
        return 1

    with tempfile.TemporaryDirectory() as td:
        zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
        zf.extractall(td)
        zf.close()

        plant_path = find_sheet(td, PLANT_GLOB)
        gen_path = find_sheet(td, GEN_GLOB)
        battery_path = None
        for g in BATTERY_GLOBS:
            battery_path = find_sheet(td, g)
            if battery_path:
                break

        if not (plant_path and gen_path):
            sys.stderr.write(
                f"SHEET_NOT_FOUND: plant={plant_path} gen={gen_path} (year={args.year})\n"
            )
            return 2

        sys.stderr.write(f"[plants]   {os.path.basename(plant_path)}\n")
        sys.stderr.write(f"[gens]     {os.path.basename(gen_path)}\n")
        sys.stderr.write(f"[battery]  {os.path.basename(battery_path) if battery_path else 'NONE'}\n")

        plant_rows = load_xlsx_rows(plant_path,
                                    ["Plant Code", "Plant Name", "State", "Latitude"])
        gen_rows = load_xlsx_rows(gen_path,
                                  ["Plant Code", "Status", "Nameplate Capacity"])

        # Filter plants to TX before joining — much smaller working set.
        state_col_p = column_lookup(plant_rows, "State")
        plant_rows_tx = [
            p for p in plant_rows
            if p.get(state_col_p) and str(p.get(state_col_p)).strip().upper() == "TX"
        ]
        sys.stderr.write(f"[tx-plants] {len(plant_rows_tx)} of {len(plant_rows)}\n")

        # Aggregate ALL generators (to avoid losing rows where State on the gen
        # row is ambiguous), then filter on the join.
        gen_agg = aggregate_generators(gen_rows)
        sys.stderr.write(f"[gen-agg]   {len(gen_agg)} plant_codes with OP units\n")

        # Build eia860_plants CSV.
        col_pc = column_lookup(plant_rows_tx, "Plant Code")
        col_name = column_lookup(plant_rows_tx, "Plant Name")
        col_lat = column_lookup(plant_rows_tx, "Latitude")
        col_lon = column_lookup(plant_rows_tx, "Longitude")
        col_county = column_lookup(plant_rows_tx, "County")
        col_sector = column_lookup(plant_rows_tx, "Sector Name", "Sector")
        col_op = column_lookup(plant_rows_tx, "Utility Name")

        plants_out = []
        for p in plant_rows_tx:
            pc = str(p.get(col_pc)).strip() if p.get(col_pc) is not None else ""
            try:
                lat = float(p.get(col_lat)) if p.get(col_lat) not in (None, "") else None
                lon = float(p.get(col_lon)) if p.get(col_lon) not in (None, "") else None
            except (TypeError, ValueError):
                lat = lon = None
            if lat is None or lon is None:
                continue
            if not (25.0 < lat < 37.0 and -107.0 < lon < -93.0):
                continue
            agg = gen_agg.get(pc, {})
            row = empty_row()
            row.update({
                "layer_id": "eia860_plants",
                "lat": lat, "lon": lon,
                "name": str(p.get(col_name) or "").strip(),
                "plant_code": pc,
                "county": str(p.get(col_county) or "").strip().upper(),
                "technology": agg.get("technology", ""),
                "sector": str(p.get(col_sector) or "").strip(),
                "fuel": agg.get("fuel", ""),
                "commissioned": agg.get("commissioned", ""),
                "capacity_mw": agg.get("capacity_mw", ""),
                "operator": str(p.get(col_op) or "").strip(),
            })
            plants_out.append(row)

        plants_path = os.path.join(args.out, f"eia860_plants_{today}.csv")
        write_csv(plants_path, plants_out)
        sys.stderr.write(f"[wrote] {plants_path} ({len(plants_out)} rows)\n")

        # Build eia860_battery CSV.
        battery_out = []
        if battery_path:
            bat_rows = load_xlsx_rows(battery_path,
                                      ["Plant Code", "Status", "Nameplate Capacity"])
            col_bpc = column_lookup(bat_rows, "Plant Code")
            col_bname = column_lookup(bat_rows, "Plant Name")
            col_bstate = column_lookup(bat_rows, "State")
            col_bcounty = column_lookup(bat_rows, "County")
            col_btech = column_lookup(bat_rows, "Storage Technology 1", "Technology")
            col_bcap = column_lookup(bat_rows, "Nameplate Capacity (MW)", "Nameplate Capacity")
            col_bop = column_lookup(bat_rows, "Utility Name")
            col_byear = column_lookup(bat_rows, "Operating Year")
            col_bstatus = column_lookup(bat_rows, "Status")

            # Coord lookup from plant sheet (battery sheet has no lat/lon).
            coord_idx = {}
            for p in plant_rows_tx:
                pc = str(p.get(col_pc)).strip() if p.get(col_pc) is not None else ""
                try:
                    coord_idx[pc] = (float(p.get(col_lat)), float(p.get(col_lon)))
                except (TypeError, ValueError):
                    pass

            for b in bat_rows:
                state = ("" if b.get(col_bstate) is None else str(b.get(col_bstate))).strip().upper()
                if state != "TX":
                    continue
                status = ("" if b.get(col_bstatus) is None else str(b.get(col_bstatus))).strip().upper()
                if status != "OP":
                    continue
                pc = str(b.get(col_bpc)).strip() if b.get(col_bpc) is not None else ""
                if pc not in coord_idx:
                    continue
                lat, lon = coord_idx[pc]
                try:
                    cap = float(b.get(col_bcap)) if b.get(col_bcap) not in (None, "") else None
                except (TypeError, ValueError):
                    cap = None
                row = empty_row()
                row.update({
                    "layer_id": "eia860_battery",
                    "lat": lat, "lon": lon,
                    "name": str(b.get(col_bname) or "").strip(),
                    "plant_code": pc,
                    "county": str(b.get(col_bcounty) or "").strip().upper(),
                    "technology": str(b.get(col_btech) or "").strip(),
                    "capacity": cap if cap is not None else "",
                    "capacity_mw": cap if cap is not None else "",
                    "operator": str(b.get(col_bop) or "").strip(),
                    "commissioned": str(b.get(col_byear) or "").strip(),
                })
                battery_out.append(row)

        battery_csv = os.path.join(args.out, f"eia860_battery_{today}.csv")
        write_csv(battery_csv, battery_out)
        sys.stderr.write(f"[wrote] {battery_csv} ({len(battery_out)} rows)\n")

    if not plants_out:
        sys.stderr.write("ERROR: eia860_plants_*.csv is empty\n")
        return 1
    print(f"OK eia860_plants={len(plants_out)} eia860_battery={len(battery_out)} src={src_url}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
