#!/usr/bin/env python3
"""Build the Hanwha-exhibit production deliverable from RRC PDQ_DSV dump.

Inputs:
  data/rrc_raw/PDQ_DSV.zip — RRC bulk PDQ export, monthly 1993-current.

  Schema (delimiter = `}`):
    GP_COUNTY_DATA_TABLE.dsv: county_no → county_name + district + fips
    OG_COUNTY_CYCLE_DATA_TABLE.dsv: county_no × cycle_year_month →
        CNTY_OIL_PROD_VOL, CNTY_GAS_PROD_VOL, CNTY_COND_PROD_VOL, CNTY_CSGD_PROD_VOL
    OG_COUNTY_LEASE_CYCLE_DATA_TABLE.dsv: county_no × lease_no × cycle_year_month
    OG_OPERATOR_CYCLE_DATA_TABLE.dsv: operator_no × cycle_year_month statewide
    OG_LEASE_CYCLE_DATA_TABLE.dsv: lease_no × cycle_year_month
    OG_REGULATORY_LEASE_DW_DATA_TABLE.dsv: lease_no → lease_name + operator_name

Outputs (NOT map layers — exported to data/ for direct distribution):
  data/production_permian6.xlsx          — 6 tabs
  data/production_permian6_summary.pdf   — 4-page summary

Scope: 6-county Permian (Pecos, Reeves, Ward, Midland, Martin, Reagan),
       1993-present, monthly aggregates.
"""
from __future__ import annotations

import argparse
import io
import os
import sys
import zipfile
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / "data" / "rrc_raw" / "PDQ_DSV.zip"
OUT_XLSX = ROOT / "data" / "production_permian6.xlsx"
OUT_PDF = ROOT / "data" / "production_permian6_summary.pdf"

SUBJECT_COUNTIES = {"PECOS", "REEVES", "WARD"}
PEER_COUNTIES = {"MIDLAND", "MARTIN", "REAGAN"}
SCOPE_COUNTIES = SUBJECT_COUNTIES | PEER_COUNTIES

DELIM = "}"


def discover_schema(zip_path: Path) -> dict:
    out = {}
    with zipfile.ZipFile(zip_path, "r") as z:
        for info in z.infolist():
            if info.is_dir():
                continue
            with z.open(info.filename) as f:
                head = f.read(8192).decode("utf-8", errors="replace")
                first_line = head.split("\n", 1)[0]
                out[info.filename] = first_line
    return out


def stream_dsv(z: zipfile.ZipFile, name: str):
    """Yield (header, row_dict) for each row in the DSV file."""
    with z.open(name) as raw:
        text = io.TextIOWrapper(raw, encoding="utf-8", errors="replace")
        header_line = text.readline().rstrip("\r\n")
        header = header_line.split(DELIM)
        for line in text:
            line = line.rstrip("\r\n")
            if not line:
                continue
            vals = line.split(DELIM)
            if len(vals) < len(header):
                vals += [""] * (len(header) - len(vals))
            elif len(vals) > len(header):
                vals = vals[: len(header)]
            yield header, dict(zip(header, vals))


def to_float(s: str) -> float:
    try:
        return float(s or 0)
    except ValueError:
        return 0.0


def to_int(s: str) -> int:
    try:
        return int(s or 0)
    except ValueError:
        return 0


def build_county_map(z: zipfile.ZipFile) -> dict:
    """Return {county_no_str: {'name': str, 'district': str}}."""
    out = {}
    for _h, row in stream_dsv(z, "GP_COUNTY_DATA_TABLE.dsv"):
        name = (row.get("COUNTY_NAME") or "").strip().upper()
        no = (row.get("COUNTY_NO") or "").strip()
        if not no:
            continue
        out[no] = {
            "name": name,
            "district": (row.get("DISTRICT_NO") or "").strip(),
            "fips": (row.get("COUNTY_FIPS_CODE") or "").strip(),
        }
    return out


def build_lease_meta(z: zipfile.ZipFile, scope_lease_keys: set) -> dict:
    """Return {(oil_gas_code, district_no, lease_no): {'lease_name', 'operator_name', 'field_name'}}.

    Only stores entries whose composite key is in scope_lease_keys."""
    out = {}
    for _h, row in stream_dsv(z, "OG_REGULATORY_LEASE_DW_DATA_TABLE.dsv"):
        og = (row.get("OIL_GAS_CODE") or "").strip()
        dist = (row.get("DISTRICT_NO") or "").strip()
        lease = (row.get("LEASE_NO") or "").strip()
        key = (og, dist, lease)
        if scope_lease_keys and key not in scope_lease_keys:
            continue
        out[key] = {
            "lease_name": (row.get("LEASE_NAME") or "").strip(),
            "operator_name": (row.get("OPERATOR_NAME") or "").strip(),
            "field_name": (row.get("FIELD_NAME") or "").strip(),
        }
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--discover", action="store_true",
                    help="Just print the zip TOC + file headers and exit")
    args = ap.parse_args()

    if not RAW.exists():
        print(f"ERROR: {RAW} not found.", file=sys.stderr)
        print("       Run `python3 scripts/fetch_pdq_dump.py` first (~3.44 GB).",
              file=sys.stderr)
        return 2

    if args.discover:
        schema = discover_schema(RAW)
        for fname, hdr in schema.items():
            print(f"=== {fname}")
            print(f"    header: {hdr[:300]}")
        return 0

    print(f"=== building Hanwha production deliverable ===")
    print(f"  source: {RAW} ({RAW.stat().st_size / (1024**3):.2f} GB)")
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(RAW, "r") as z:
        print("  [1/6] county_no map...")
        county_map = build_county_map(z)
        scope_county_nos = {
            no for no, meta in county_map.items()
            if meta["name"] in SCOPE_COUNTIES
        }
        print(f"        scope: {sorted(meta['name'] for no, meta in county_map.items() if no in scope_county_nos)}")
        if not scope_county_nos:
            print("ERROR: scope counties not found in GP_COUNTY_DATA_TABLE",
                  file=sys.stderr)
            return 3

        print("  [2/6] streaming OG_COUNTY_CYCLE_DATA_TABLE (county × month)...")
        county_monthly: dict[tuple[str, str], dict] = {}
        n = 0
        for _h, row in stream_dsv(z, "OG_COUNTY_CYCLE_DATA_TABLE.dsv"):
            n += 1
            cno = (row.get("COUNTY_NO") or "").strip()
            if cno not in scope_county_nos:
                continue
            ym = (row.get("CYCLE_YEAR_MONTH") or "").strip()
            if not ym or len(ym) < 6:
                continue
            agg = county_monthly.setdefault((cno, ym), {
                "oil_bbl": 0.0, "gas_mcf": 0.0, "cond_bbl": 0.0, "csgd_mcf": 0.0,
            })
            agg["oil_bbl"] += to_float(row.get("CNTY_OIL_PROD_VOL"))
            agg["gas_mcf"] += to_float(row.get("CNTY_GAS_PROD_VOL"))
            agg["cond_bbl"] += to_float(row.get("CNTY_COND_PROD_VOL"))
            agg["csgd_mcf"] += to_float(row.get("CNTY_CSGD_PROD_VOL"))
        print(f"        rows scanned: {n:,}  scope month-county cells: {len(county_monthly):,}")

        print("  [3/6] streaming OG_COUNTY_LEASE_CYCLE_DATA_TABLE for top leases/operators...")
        # Aggregate by (oil_gas, district, lease, operator) across in-scope months.
        lease_agg: dict[tuple[str, str, str, str], dict] = {}
        operator_agg: dict[tuple[str, str], dict] = {}  # (operator_no, county_no) -> agg
        n2 = 0
        for _h, row in stream_dsv(z, "OG_COUNTY_LEASE_CYCLE_DATA_TABLE.dsv"):
            n2 += 1
            cno = (row.get("COUNTY_NO") or "").strip()
            if cno not in scope_county_nos:
                continue
            ym = (row.get("CYCLE_YEAR_MONTH") or "").strip()
            if not ym or len(ym) < 6:
                continue
            og = (row.get("OIL_GAS_CODE") or "").strip()
            dist = (row.get("DISTRICT_NO") or "").strip()
            lease = (row.get("LEASE_NO") or "").strip()
            op = (row.get("OPERATOR_NO") or "").strip()
            lkey = (og, dist, lease, op)
            la = lease_agg.setdefault(lkey, {
                "county_no": cno, "oil_bbl": 0.0, "gas_mcf": 0.0,
                "cond_bbl": 0.0, "csgd_mcf": 0.0,
                "first_ym": ym, "last_ym": ym,
            })
            la["oil_bbl"] += to_float(row.get("CNTY_LSE_OIL_PROD_VOL"))
            la["gas_mcf"] += to_float(row.get("CNTY_LSE_GAS_PROD_VOL"))
            la["cond_bbl"] += to_float(row.get("CNTY_LSE_COND_PROD_VOL"))
            la["csgd_mcf"] += to_float(row.get("CNTY_LSE_CSGD_PROD_VOL"))
            if ym < la["first_ym"]:
                la["first_ym"] = ym
            if ym > la["last_ym"]:
                la["last_ym"] = ym
            okey = (op, cno)
            oa = operator_agg.setdefault(okey, {"oil_bbl": 0.0, "gas_mcf": 0.0})
            oa["oil_bbl"] += to_float(row.get("CNTY_LSE_OIL_PROD_VOL"))
            oa["gas_mcf"] += to_float(row.get("CNTY_LSE_GAS_PROD_VOL"))
        print(f"        lease-county-month rows scanned: {n2:,}")
        print(f"        unique in-scope leases: {len(lease_agg):,}")
        print(f"        unique in-scope (operator, county) pairs: {len(operator_agg):,}")

        print("  [4/6] lease metadata (lease_no → lease_name + operator_name + field_name)...")
        lease_keys_3 = {(og, dist, lease) for (og, dist, lease, _op) in lease_agg}
        lease_meta = build_lease_meta(z, lease_keys_3)
        print(f"        lease metadata rows matched: {len(lease_meta):,}")

    # --------- Aggregation helpers ----------
    def ym_to_year(ym: str) -> int:
        try:
            return int(ym[:4])
        except ValueError:
            return 0

    # Tab 1: annual by county
    annual_by_county: dict[tuple[int, str], dict] = {}  # (year, county_name) -> agg
    monthly_total: dict[str, dict] = {}  # ym -> agg (6-county totals)
    sale_vs_peer: dict[tuple[int, str], dict] = {}  # (year, role) -> agg
    for (cno, ym), agg in county_monthly.items():
        y = ym_to_year(ym)
        if y == 0:
            continue
        cname = county_map[cno]["name"]
        role = "subject" if cname in SUBJECT_COUNTIES else "peer"
        ka = annual_by_county.setdefault((y, cname), {
            "oil_bbl": 0.0, "gas_mcf": 0.0, "cond_bbl": 0.0, "csgd_mcf": 0.0,
            "role": role,
        })
        for k in ("oil_bbl", "gas_mcf", "cond_bbl", "csgd_mcf"):
            ka[k] += agg[k]
        km = monthly_total.setdefault(ym, {
            "oil_bbl": 0.0, "gas_mcf": 0.0, "cond_bbl": 0.0, "csgd_mcf": 0.0,
        })
        for k in ("oil_bbl", "gas_mcf", "cond_bbl", "csgd_mcf"):
            km[k] += agg[k]
        ks = sale_vs_peer.setdefault((y, role), {
            "oil_bbl": 0.0, "gas_mcf": 0.0, "cond_bbl": 0.0, "csgd_mcf": 0.0,
            "counties": set(),
        })
        for k in ("oil_bbl", "gas_mcf", "cond_bbl", "csgd_mcf"):
            ks[k] += agg[k]
        ks["counties"].add(cname)

    # Tab 2: top operators (statewide-in-scope by aggregate oil + gas-as-boe)
    op_total: dict[str, dict] = defaultdict(lambda: {"oil_bbl": 0.0, "gas_mcf": 0.0})
    for (op, cno), agg in operator_agg.items():
        op_total[op]["oil_bbl"] += agg["oil_bbl"]
        op_total[op]["gas_mcf"] += agg["gas_mcf"]
    # Get operator names — re-stream operator DW table
    print("  [5/6] operator name lookup...")
    op_names: dict[str, str] = {}
    with zipfile.ZipFile(RAW, "r") as z:
        for _h, row in stream_dsv(z, "OG_OPERATOR_DW_DATA_TABLE.dsv"):
            op = (row.get("OPERATOR_NO") or "").strip()
            if op in op_total:
                op_names[op] = (row.get("OPERATOR_NAME") or "").strip()

    # BOE for ranking: 1 BOE = 1 bbl oil = 6 mcf gas. Use boe-equivalent for rank.
    def boe(d: dict) -> float:
        return d["oil_bbl"] + d["gas_mcf"] / 6.0
    top_ops = sorted(op_total.items(), key=lambda kv: -boe(kv[1]))[:20]

    # Tab 3: top 100 leases by aggregate BOE
    lease_total = []
    for (og, dist, lease, op), agg in lease_agg.items():
        b = boe(agg)
        meta = lease_meta.get((og, dist, lease), {})
        lease_total.append({
            "og": og, "district": dist, "lease": lease, "operator_no": op,
            "lease_name": meta.get("lease_name", ""),
            "operator_name": meta.get("operator_name", "") or op_names.get(op, ""),
            "field_name": meta.get("field_name", ""),
            "county_no": agg["county_no"],
            "county_name": county_map[agg["county_no"]]["name"],
            "oil_bbl": agg["oil_bbl"],
            "gas_mcf": agg["gas_mcf"],
            "boe": b,
            "first_ym": agg["first_ym"],
            "last_ym": agg["last_ym"],
        })
    lease_total.sort(key=lambda r: -r["boe"])
    top_leases = lease_total[:100]

    # --------- Write Excel ----------
    print("  [6/6] writing Excel + PDF...")
    import openpyxl
    from openpyxl import Workbook
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Annual by county"
    ws1.append(["Year", "County", "Role", "Oil (bbl)", "Gas (mcf)", "Condensate (bbl)", "Casinghead gas (mcf)"])
    for (y, c), agg in sorted(annual_by_county.items()):
        ws1.append([y, c, agg["role"],
                    round(agg["oil_bbl"]), round(agg["gas_mcf"]),
                    round(agg["cond_bbl"]), round(agg["csgd_mcf"])])

    ws2 = wb.create_sheet("Top 20 operators")
    ws2.append(["Rank", "Operator No", "Operator Name", "Oil (bbl)", "Gas (mcf)", "BOE"])
    for i, (op, agg) in enumerate(top_ops, 1):
        ws2.append([i, op, op_names.get(op, ""),
                    round(agg["oil_bbl"]), round(agg["gas_mcf"]), round(boe(agg))])

    ws3 = wb.create_sheet("Top 100 leases")
    ws3.append(["Rank", "County", "District", "Lease No", "Lease Name", "Operator",
                "Field", "Oil (bbl)", "Gas (mcf)", "BOE", "First month", "Last month"])
    for i, lr in enumerate(top_leases, 1):
        ws3.append([i, lr["county_name"], lr["district"], lr["lease"], lr["lease_name"],
                    lr["operator_name"], lr["field_name"],
                    round(lr["oil_bbl"]), round(lr["gas_mcf"]), round(lr["boe"]),
                    lr["first_ym"], lr["last_ym"]])

    ws4 = wb.create_sheet("Monthly time-series")
    ws4.append(["Year-Month", "Oil (bbl)", "Gas (mcf)", "Condensate (bbl)", "Casinghead gas (mcf)"])
    for ym, agg in sorted(monthly_total.items()):
        ws4.append([ym,
                    round(agg["oil_bbl"]), round(agg["gas_mcf"]),
                    round(agg["cond_bbl"]), round(agg["csgd_mcf"])])

    ws5 = wb.create_sheet("Sale-area vs peer")
    ws5.append(["Year", "Subject oil (bbl)", "Subject gas (mcf)",
                "Peer oil (bbl)", "Peer gas (mcf)",
                "Peer/subject oil ratio", "Peer/subject gas ratio",
                "Subject per-county oil/yr", "Peer per-county oil/yr"])
    years = sorted({y for (y, _r) in sale_vs_peer})
    for y in years:
        s = sale_vs_peer.get((y, "subject"), {"oil_bbl": 0.0, "gas_mcf": 0.0, "counties": set()})
        p = sale_vs_peer.get((y, "peer"), {"oil_bbl": 0.0, "gas_mcf": 0.0, "counties": set()})
        s_cnt = max(len(s["counties"]), 1)
        p_cnt = max(len(p["counties"]), 1)
        oil_ratio = (p["oil_bbl"] / s["oil_bbl"]) if s["oil_bbl"] > 0 else 0
        gas_ratio = (p["gas_mcf"] / s["gas_mcf"]) if s["gas_mcf"] > 0 else 0
        ws5.append([y,
                    round(s["oil_bbl"]), round(s["gas_mcf"]),
                    round(p["oil_bbl"]), round(p["gas_mcf"]),
                    round(oil_ratio, 2), round(gas_ratio, 2),
                    round(s["oil_bbl"] / s_cnt), round(p["oil_bbl"] / p_cnt)])

    ws6 = wb.create_sheet("Raw — county-month")
    ws6.append(["Year-Month", "County", "Role", "Oil (bbl)", "Gas (mcf)", "Condensate (bbl)", "Casinghead gas (mcf)"])
    for (cno, ym), agg in sorted(county_monthly.items()):
        cname = county_map[cno]["name"]
        role = "subject" if cname in SUBJECT_COUNTIES else "peer"
        ws6.append([ym, cname, role,
                    round(agg["oil_bbl"]), round(agg["gas_mcf"]),
                    round(agg["cond_bbl"]), round(agg["csgd_mcf"])])

    tmp_xlsx = OUT_XLSX.with_suffix(".xlsx.tmp")
    wb.save(tmp_xlsx)
    os.replace(tmp_xlsx, OUT_XLSX)
    print(f"        wrote {OUT_XLSX} ({OUT_XLSX.stat().st_size / 1024:.1f} KB)")

    # --------- Write PDF ----------
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages

    # Subject vs peer headline
    yrs = sorted({y for (y, _r) in sale_vs_peer})
    s_oil = [sale_vs_peer.get((y, "subject"), {"oil_bbl": 0.0})["oil_bbl"] for y in yrs]
    p_oil = [sale_vs_peer.get((y, "peer"), {"oil_bbl": 0.0})["oil_bbl"] for y in yrs]
    s_gas = [sale_vs_peer.get((y, "subject"), {"gas_mcf": 0.0})["gas_mcf"] for y in yrs]
    p_gas = [sale_vs_peer.get((y, "peer"), {"gas_mcf": 0.0})["gas_mcf"] for y in yrs]

    tmp_pdf = OUT_PDF.with_suffix(".pdf.tmp")
    with PdfPages(tmp_pdf) as pdf:
        # Page 1: Headline
        fig, axes = plt.subplots(2, 1, figsize=(11, 8.5))
        ax = axes[0]
        ax.plot(yrs, [v / 1e6 for v in s_oil], label="Sale area (Pecos+Reeves+Ward)", color="#7f1d1d", lw=2)
        ax.plot(yrs, [v / 1e6 for v in p_oil], label="Peer (Midland+Martin+Reagan)", color="#0f766e", lw=2)
        ax.set_title("Annual oil production — sale area vs peer (RRC PDQ 1993-present)")
        ax.set_ylabel("Million bbl / year")
        ax.set_xlabel("Year")
        ax.legend(loc="upper left")
        ax.grid(True, alpha=0.3)

        ax = axes[1]
        ax.plot(yrs, [v / 1e6 for v in s_gas], label="Sale area gas", color="#7f1d1d", lw=2, linestyle="--")
        ax.plot(yrs, [v / 1e6 for v in p_gas], label="Peer gas", color="#0f766e", lw=2, linestyle="--")
        ax.set_title("Annual gas production — sale area vs peer")
        ax.set_ylabel("Million mcf / year")
        ax.set_xlabel("Year")
        ax.legend(loc="upper left")
        ax.grid(True, alpha=0.3)
        plt.tight_layout()
        pdf.savefig(fig)
        plt.close(fig)

        # Page 2: per-county small multiples
        fig, axes = plt.subplots(2, 3, figsize=(11, 8.5))
        ordered = ["PECOS", "REEVES", "WARD", "MIDLAND", "MARTIN", "REAGAN"]
        for i, cname in enumerate(ordered):
            ax = axes[i // 3][i % 3]
            ax.plot(yrs, [annual_by_county.get((y, cname), {"oil_bbl": 0.0})["oil_bbl"] / 1e6 for y in yrs],
                    color="#7f1d1d" if cname in SUBJECT_COUNTIES else "#0f766e", lw=2)
            role = "SUBJECT" if cname in SUBJECT_COUNTIES else "PEER"
            ax.set_title(f"{cname} ({role}) — oil (Mbbl/yr)")
            ax.grid(True, alpha=0.3)
            ax.set_xlabel("Year")
        plt.tight_layout()
        pdf.savefig(fig)
        plt.close(fig)

        # Page 3: Top operators (text)
        fig = plt.figure(figsize=(11, 8.5))
        plt.axis("off")
        lines = ["Top 20 operators by aggregate BOE — RRC PDQ 1993-present, 6-county Permian", ""]
        for i, (op, agg) in enumerate(top_ops, 1):
            name = op_names.get(op, "(no name)")[:55]
            lines.append(f"  {i:2}.  {name:55}  oil {agg['oil_bbl']/1e6:8.2f} Mbbl   gas {agg['gas_mcf']/1e6:8.2f} Mmcf")
        plt.text(0.05, 0.95, "\n".join(lines), family="monospace", size=8, va="top")
        pdf.savefig(fig)
        plt.close(fig)

        # Page 4: Methodology
        fig = plt.figure(figsize=(11, 8.5))
        plt.axis("off")
        from datetime import date
        methodology = [
            "Methodology — Hanwha-exhibit production deliverable",
            "",
            "Source: RRC Production Data Query (PDQ) bulk DSV dump.",
            "  URL: https://mft.rrc.texas.gov/link/1f5ddb8d-329a-4459-b7f8-177b4f5ee60d",
            f"  Snapshot date: {RAW.stat().st_mtime}",
            "  Update cadence: last Saturday of each month",
            "",
            "Scope:",
            "  Subject counties (sale area): Pecos, Reeves, Ward",
            "  Active Permian peer:          Midland, Martin, Reagan",
            "  Date range: 1993-01 through latest PDQ cycle.",
            "",
            "Aggregation:",
            "  County-level monthly production summed across all leases per",
            "  county (OG_COUNTY_CYCLE_DATA_TABLE.dsv). Lease-level totals",
            "  via OG_COUNTY_LEASE_CYCLE_DATA_TABLE.dsv. Operator-level",
            "  totals via OG_OPERATOR_DW_DATA_TABLE.dsv lookups.",
            "",
            "Notes:",
            "  - Oil + gas are reported volumes (bbl, mcf). No price adjustment.",
            "  - BOE = oil_bbl + gas_mcf/6 (standard energy-equivalent ratio).",
            "  - Per-county-per-year normalization uses the actual county count",
            "    per side (3 subject / 3 peer).",
            "  - Cyclical month-level seasonality not removed.",
            "",
            f"Generated: {date.today().isoformat()} by scripts/build_production_deliverable.py",
        ]
        plt.text(0.05, 0.95, "\n".join(methodology), family="monospace", size=9, va="top")
        pdf.savefig(fig)
        plt.close(fig)

    os.replace(tmp_pdf, OUT_PDF)
    print(f"        wrote {OUT_PDF} ({OUT_PDF.stat().st_size / 1024:.1f} KB)")

    print("\n=== done ===")
    return 0


if __name__ == "__main__":
    sys.exit(main())
