#!/usr/bin/env python3
"""Build the StreetEasy listings xlsx from listings.json (+ research_fill.json)."""
import json, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SCRATCH = os.environ.get("SE_WORKDIR", os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + "/data")
OUT = sys.argv[1] if len(sys.argv) > 1 else os.path.join(SCRATCH, "streeteasy_listings_2026-07-12.xlsx")

rows = json.load(open(os.path.join(SCRATCH, "listings.json")))
fill_path = os.path.join(SCRATCH, "research_fill.json")
fills = json.load(open(fill_path)) if os.path.exists(fill_path) else {}
for r in rows:
    f = fills.get(r["listing_id"])
    if not f:
        continue
    if f.get("sqft") and not r.get("sqft"):
        r["sqft"] = f["sqft"]
        r["sqft_source"] = f.get("source", "research")
    if f.get("year_built") and not r.get("year_built"):
        r["year_built"] = f["year_built"]

rows.sort(key=lambda r: (r.get("neighborhood") or "", r.get("rent") or 0))

def j(v, sep="; "):
    if isinstance(v, list):
        return sep.join(str(x) for x in v) if v else ""
    return "" if v is None else v

def hist(r):
    return "; ".join(f"{h['date']}: ${h['price']:,} {h['status']}" for h in r.get("unit_history") or [] if h.get("date"))

def pchanges(r):
    return "; ".join(f"{c['date']}: ${c['price']:,}" for c in r.get("price_changes") or [] if c.get("price"))

COLS = [
    ("Listing ID", lambda r: r["listing_id"], 10),
    ("URL", lambda r: r["url"], 34),
    ("Address", lambda r: j(r.get("full_address")), 26),
    ("Unit", lambda r: j(r.get("unit")), 9),
    ("Neighborhood", lambda r: j(r.get("neighborhood")), 15),
    ("Borough", lambda r: j(r.get("borough")), 10),
    ("Zip", lambda r: j(r.get("zip")), 7),
    ("Rent ($/mo)", lambda r: r.get("rent"), 11),
    ("Net Effective Rent", lambda r: r.get("net_effective_rent"), 12),
    ("Months Free", lambda r: r.get("months_free"), 8),
    ("Lease Term (mo)", lambda r: r.get("lease_term_months"), 9),
    ("Beds", lambda r: r.get("bedrooms"), 6),
    ("Baths", lambda r: r.get("baths"), 6),
    ("Half Baths", lambda r: r.get("half_baths"), 7),
    ("Rooms", lambda r: r.get("rooms"), 7),
    ("Sq Ft", lambda r: r.get("sqft") or "", 8),
    ("$/SqFt/Yr", lambda r: round(r["rent"] * 12 / r["sqft"], 0) if r.get("sqft") and r.get("rent") else "", 9),
    ("Unit Type", lambda r: j(r.get("unit_type")), 13),
    ("Status", lambda r: j(r.get("status")), 9),
    ("Listed Date", lambda r: j(r.get("listed_date")), 11),
    ("Days on Market", lambda r: r.get("days_on_market"), 9),
    ("Available", lambda r: j(r.get("available_date")), 11),
    ("Private Outdoor Space", lambda r: j(r.get("private_outdoor_space")), 16),
    ("Views", lambda r: j(r.get("views")), 10),
    ("Unit Features", lambda r: j(r.get("unit_features")), 40),
    ("Building Amenities", lambda r: j(r.get("building_amenities")), 50),
    ("Doorman", lambda r: j(r.get("doorman_types")), 10),
    ("Parking", lambda r: j(r.get("parking_types")), 12),
    ("Shared Outdoor", lambda r: j(r.get("shared_outdoor_space")), 16),
    ("Storage", lambda r: j(r.get("storage_types")), 12),
    ("Building Name", lambda r: j(r.get("building_name")), 16),
    ("Year Built", lambda r: r.get("year_built") or "", 8),
    ("Floors", lambda r: r.get("building_floors"), 7),
    ("Units in Bldg", lambda r: r.get("building_units"), 8),
    ("Brokerage", lambda r: j(r.get("brokerage")), 18),
    ("Broker License Name", lambda r: j(r.get("broker_license_name")), 18),
    ("Broker License Type", lambda r: j(r.get("broker_license_type")), 18),
    ("Broker Address", lambda r: j(r.get("broker_address")), 26),
    ("Broker Fee", lambda r: j(r.get("broker_fee")) or "None listed", 12),
    ("Security Deposit", lambda r: r.get("security_deposit"), 11),
    ("Fees", lambda r: j(r.get("fees")), 45),
    ("Price History", lambda r: pchanges(r), 24),
    ("Unit Rental History", lambda r: hist(r), 45),
    ("Area Median Rent (same beds)", lambda r: r.get("area_median_rent_same_beds"), 12),
    ("Open Houses", lambda r: j(r.get("open_houses")), 14),
    ("Photos", lambda r: r.get("photo_count"), 7),
    ("Videos", lambda r: r.get("video_count"), 7),
    ("Floorplans", lambda r: r.get("floorplan_count"), 7),
    ("3D Tour", lambda r: j(r.get("tour_3d_url")), 12),
    ("MLS #", lambda r: j(r.get("mls_number")), 9),
    ("Listing Source", lambda r: j(r.get("listing_source_type")), 11),
    ("Latitude", lambda r: r.get("latitude"), 10),
    ("Longitude", lambda r: r.get("longitude"), 10),
    ("SqFt Source", lambda r: j(r.get("sqft_source")), 24),
    ("Description", lambda r: j(r.get("description")), 80),
]

wb = Workbook()
ws = wb.active
ws.title = "Listings 2026-07-12"

hdr_fill = PatternFill("solid", fgColor="1F3864")
hdr_font = Font(color="FFFFFF", bold=True, size=10)
thin = Border(bottom=Side(style="thin", color="D0D0D0"))

for c, (name, _, width) in enumerate(COLS, 1):
    cell = ws.cell(row=1, column=c, value=name)
    cell.fill = hdr_fill; cell.font = hdr_font
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(c)].width = width
ws.freeze_panes = "C2"
ws.row_dimensions[1].height = 28

alt = PatternFill("solid", fgColor="F2F6FC")
for i, r in enumerate(rows, 2):
    for c, (name, fn, _) in enumerate(COLS, 1):
        v = fn(r)
        cell = ws.cell(row=i, column=c, value=v)
        cell.border = thin
        cell.alignment = Alignment(vertical="top", wrap_text=name in ("Description", "Unit Features", "Building Amenities", "Fees", "Unit Rental History"))
        if i % 2 == 0:
            cell.fill = alt
        if name == "URL":
            cell.hyperlink = v; cell.font = Font(color="0563C1", underline="single", size=10)
        elif name in ("Rent ($/mo)", "Net Effective Rent", "Security Deposit", "Area Median Rent (same beds)"):
            cell.number_format = '$#,##0'
        elif name == "$/SqFt/Yr":
            cell.number_format = '$#,##0'
        else:
            cell.font = Font(size=10)
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(rows)+1}"

# Summary sheet
s = wb.create_sheet("Summary")
s.column_dimensions['A'].width = 30; s.column_dimensions['B'].width = 18
rents = [r["rent"] for r in rows if r.get("rent")]
def put(row, a, b, bold=False, fmt=None):
    ca = s.cell(row=row, column=1, value=a); cb = s.cell(row=row, column=2, value=b)
    if bold: ca.font = Font(bold=True); cb.font = Font(bold=True)
    if fmt: cb.number_format = fmt
put(1, f"StreetEasy Listings — shared 7/12–7/14/2026 ({len(rows)} listings)", "", bold=True)
put(3, "Listings", len(rows))
put(4, "Median rent", sorted(rents)[len(rents)//2], fmt='$#,##0')
put(5, "Min rent", min(rents), fmt='$#,##0')
put(6, "Max rent", max(rents), fmt='$#,##0')
row = 8
put(row, "By neighborhood", "", bold=True); row += 1
from collections import Counter
for hood, n in Counter(r["neighborhood"] for r in rows).most_common():
    put(row, hood, n); row += 1
row += 1
put(row, "By bedrooms", "", bold=True); row += 1
for bd, n in sorted(Counter(r["bedrooms"] for r in rows).items()):
    put(row, f"{bd} BR", n); row += 1

wb.save(OUT)
print("wrote", OUT, len(rows), "rows,", len(COLS), "columns")
